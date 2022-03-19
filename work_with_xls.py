import openpyxl
import os

file = os.path.join(os.getcwd(), 'base.xlsx')


file_excel = openpyxl.load_workbook(file)
list_excel = file_excel['list']



list_of_genres = []
list_of_singer = []
for i in range(2, list_excel.max_row + 1):
    name_singer = list_excel.cell(column=3, row=i).value
    name_of_genre = list_excel.cell(column=4, row=i).value
    list_of_singer.append(name_singer)
    list_of_genres.append(name_of_genre)













