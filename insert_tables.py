from main import connect_one
import openpyxl
from work_with_xls import list_of_singer, list_of_genres, file
from transliterate import translit


def insert_genres():
    for i in list_of_genres:
        connect_one.connect().execute(f"insert into list_of_genres (genres) values  ('{i}') "
                                      f"ON CONFLICT DO NOTHING")


def insert_artists():
    for i in list_of_singer:
        connect_one.connect().execute(f"INSERT INTO list_of_singers (singer_name) VALUES ('{i}')"
                                      f"ON CONFLICT DO NOTHING")


def insert_genres_artists():
    file_excel = openpyxl.load_workbook(file)
    list_excel = file_excel['list']
    list_id_singer = []
    for i in range(2, list_excel.max_row + 1):
        name_singer = list_excel.cell(row=i, column=3).value
        tuple_id_singer = connect_one.connect().execute(
            f"SELECT id FROM list_of_singers WHERE singer_name = '{name_singer}';").fetchone()
        for k in tuple_id_singer:
            list_id_singer.append(k)

    for i, v in zip(range(2, list_excel.max_row + 1), list_id_singer):
        list_excel.cell(row=i, column=2).value = v
    file_excel.save(file)

    list_id_genres = []
    for i in range(2, list_excel.max_row + 1):
        name_genres = list_excel.cell(row=i, column=4).value
        tuple_ids_genres = connect_one.connect().execute(
            f"SELECT id FROM list_of_genres WHERE genres = '{name_genres}';").fetchone()
        for k in tuple_ids_genres:
            list_id_genres.append(k)

    for i, v in zip(range(2, list_excel.max_row + 1), list_id_genres):
        list_excel.cell(row=i, column=4).value = v
        file_excel.save(file)

    for i, v in zip(list_id_genres, list_id_singer):
        connect_one.connect().execute(f"INSERT INTO singers_genres (id_singer, id_genres) VALUES ('{v}', '{i}')"
                                      f"ON CONFLICT DO NOTHING;")


def insert_singer_album():
    file_excel = openpyxl.load_workbook(file)
    list_excel_albums = file_excel['albums']
    list_excel = file_excel['list']
    list_of_names = []

    for i in range(2, list_excel.max_row + 1):
        count = 0
        names = list_excel.cell(column=3, row=i).value
        while count != 3:
            list_of_names.append(names)
            count = count + 1

    for i in list_of_names:
        while list_of_names.count(i) > 3:
            list_of_names.remove(i)

    count = 2
    m = 1
    for i in list_of_names:
        list_excel_albums.cell(column=3, row=count).value = i
        list_excel_albums.cell(column=1, row=count).value = m
        m = m + 1
        count = count + 1
        file_excel.save(file)
    count = 2
    for i in list_of_names:
        ids_tuple = connect_one.connect().execute(
            f"SELECT id FROM list_of_singers WHERE singer_name = '{i}'").fetchone()
        for k in ids_tuple:
            list_excel_albums.cell(column=2, row=count).value = k
            count = count + 1
            file_excel.save(file)

    for i in range(2, list_excel_albums.max_row + 1):
        name_of_album = list_excel_albums.cell(column=5, row=i).value
        year = list_excel_albums.cell(column=6, row=i).value
        connect_one.connect().execute(f"INSERT INTO album_list (album_name, year) VALUES ('{name_of_album}', '{year}')")

    for i in range(2, list_excel_albums.max_row + 1):
        name_of_album = list_excel_albums.cell(column=5, row=i).value
        id_tuple = connect_one.connect().execute(
            f"SELECT id FROM album_list WHERE album_name = '{name_of_album}'").fetchone()
        for k in id_tuple:
            list_excel_albums.cell(column=4, row=i).value = k
            file_excel.save(file)

    for i in range(2, list_excel_albums.max_row + 1):
        id_album = list_excel_albums.cell(column=4, row=i).value
        id_singer = list_excel_albums.cell(column=2, row=i).value
        connect_one.connect().execute(
            f"INSERT INTO singer_album (id_singer, id_album) VALUES ('{id_singer}', '{id_album}')"
            f"ON CONFLICT DO NOTHING;")


def insert_trucks():
    file_excel = openpyxl.load_workbook(file)
    list_excel = file_excel['trucks_lists']

    list_album = []
    tuple_album = connect_one.connect().execute(f"SELECT album_name FROM album_list").fetchall()
    for i in tuple_album:
        for k in i:
            list_album.append(k)

    count = 1
    count_one = 2
    for i in list_album:
        count_two = 0
        while count_two < 3:
            list_excel.cell(column=3, row=count_one).value = i
            list_excel.cell(column=1, row=count_one).value = count
            count_two = count_two + 1
            count_one = count_one + 1
            count = count + 1
            file_excel.save(file)

    for i in range(2, list_excel.max_row + 1):
        name_of_album = list_excel.cell(column=3, row=i).value
        tuple_id_album = connect_one.connect().execute(
            f"SELECT id FROM album_list WHERE album_name = '{name_of_album}'").fetchone()
        for k in tuple_id_album:
            list_excel.cell(column=2, row=i).value = k
            file_excel.save(file)

    for i in range(2, list_excel.max_row + 1):
        name_of_tracks = list_excel.cell(column=5, row=i).value
        id_albums = list_excel.cell(column=2, row=i).value
        duration_track = list_excel.cell(column=6, row=i).value
        connect_one.connect().execute(
            f"INSERT INTO tracks (name_of_the_track, album_id, duration) VALUES ('{name_of_tracks}', '{id_albums}', '{duration_track}')"
            )
        tuple_id_truck = connect_one.connect().execute(
            f"SELECT id from tracks WHERE name_of_the_track = '{name_of_tracks}'").fetchone()
        for k in tuple_id_truck:
            list_excel.cell(column=4, row=i).value = k
            file_excel.save(file)

    for i in range(2, list_excel.max_row + 1):
        id_tracks = list_excel.cell(column=4, row=i).value
        tuple_id_album = connect_one.connect().execute(
            f"SELECT album_id from tracks WHERE id = '{id_tracks}'").fetchone()
        tuple_name_truck = connect_one.connect().execute(
            f"SELECT name_of_the_track from tracks WHERE id = '{id_tracks}'").fetchone()
        for id_album_from_trucks, name_truck in zip(tuple_id_album, tuple_name_truck):
            name_truck_eng = (translit(name_truck, 'ru', reversed=True)).replace(' ', '_').replace('»', '').replace('(',
                                                                                                                    '').replace(
                ')', '').replace('’', '').replace("'", '').replace("\"", '').strip()
            tuple_ids_albums = connect_one.connect().execute(
                f"SELECT id from album_list where id = '{id_album_from_trucks}'").fetchone()
            for k in tuple_ids_albums:
                tuple_id_singer = connect_one.connect().execute(
                    f"SELECT id_singer FROM singer_album WHERE id_album = '{k}'").fetchone()
                for id_singer in tuple_id_singer:
                    tuple_name_singer = connect_one.connect().execute(
                        f"SELECT singer_name FROM list_of_singers WHERE id = '{id_singer}'").fetchone()
                    for name_singer in tuple_name_singer:
                        name_singer_eng = (translit(name_singer, 'ru', reversed=True)).replace(' ', '_').replace('»',
                                                                                                                 '').replace(
                            '(', '').replace(')', '').replace('’', '').replace("'", '').replace("\"", '').strip()
                        link = f'https://my_site/download/{name_singer_eng}_{name_truck_eng}.mp3'
                        connect_one.connect().execute(
                            f"UPDATE tracks SET track_link = '{link}' WHERE name_of_the_track = '{name_truck}'")
                        list_excel.cell(column=7, row=i).value = link
                        file_excel.save(file)


def insert_compilation():
    set_compilation = set()
    file_excel = openpyxl.load_workbook(file)
    list_excel = file_excel['sbornik']
    for i in range(2, list_excel.max_row + 1):
        name_track = list_excel.cell(column=3, row=i).value
        id_track_tuple = connect_one.connect().execute(
            f"SELECT id FROM tracks WHERE name_of_the_track='{name_track}'").fetchone()
        set_compilation.add(list_excel.cell(column=5, row=i).value.strip())
        for id_track in id_track_tuple:
            list_excel.cell(column=2, row=i).value = id_track
            file_excel.save(file)

    for i in set_compilation:
        connect_one.connect().execute(f"INSERT INTO compilation (name) VALUES ('{i}')"
                                      f"ON CONFLICT DO NOTHING")

    for i in range(2, list_excel.max_row + 1):
        name_compilation = list_excel.cell(column=5, row=i).value
        year_compilation = list_excel.cell(column=6, row=i).value
        connect_one.connect().execute(
            f"UPDATE compilation SET year = '{year_compilation}' WHERE name = '{name_compilation}'")
        tuple_id_compilation = connect_one.connect().execute(
            f"SELECT id FROM compilation WHERE name = '{name_compilation}'").fetchone()
        for id_compilation in tuple_id_compilation:
            list_excel.cell(column=4, row=i).value = id_compilation
            file_excel.save(file)
        id_track = list_excel.cell(column=2, row=i).value
        id_compilation = list_excel.cell(column=4, row=i).value
        connect_one.connect().execute(
            f"INSERT INTO compilation_track (id_track, id_compilation) VALUES ('{id_track}', '{id_compilation}')"
            f"ON CONFLICT DO NOTHING")

insert_artists()
insert_genres()
insert_genres_artists()
insert_singer_album()
insert_trucks()
insert_compilation()
